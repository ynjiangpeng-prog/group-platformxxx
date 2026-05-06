import io
from datetime import datetime


_STATUS_MAP = {
    "充电完成": "completed",
    "充电中": "charging",
    "已结束": "completed",
    "待结算": "pending",
    "已取消": "cancelled",
}

_PAY_STATUS_MAP = {
    "已支付": "paid",
    "未支付": "unpaid",
    "已退款": "refunded",
}


class ChargingOrderImportService:

    @staticmethod
    def build_header_index(headers: list[str | None]) -> dict[str, int]:
        result = {}
        for idx, h in enumerate(headers):
            if h is None:
                continue
            h = str(h).strip()
            if h:
                result[h] = idx
        return result

    @staticmethod
    def parse_datetime(text: str) -> datetime | None:
        if not text:
            return None
        text = str(text).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def parse_xlsx(file_data: bytes) -> list[dict]:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        sheet = wb.active

        rows_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            wb.close()
            return []

        header_idx = ChargingOrderImportService.build_header_index(list(header_row))

        def _get(row, col_name, default=None):
            col = header_idx.get(col_name)
            if col is None or col >= len(row):
                return default
            val = row[col]
            if val is None:
                return default
            return val

        def _str(row, col_name):
            val = _get(row, col_name)
            if val is None:
                return None
            return str(val).strip() or None

        def _float(row, col_name):
            val = _get(row, col_name)
            if val is None:
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        def _int(row, col_name):
            val = _get(row, col_name)
            if val is None:
                return None
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return None

        results = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            source_order_no = _str(row, "平台订单号")
            if not source_order_no:
                continue

            raw_status = _str(row, "充电状态") or ""
            status = _STATUS_MAP.get(raw_status, raw_status.lower() if raw_status else "completed")

            raw_pay_status = _str(row, "支付状态") or ""
            pay_status = _PAY_STATUS_MAP.get(raw_pay_status, raw_pay_status.lower() if raw_pay_status else "unpaid")

            start_soc = _int(row, "初始SOC")
            end_soc = _int(row, "结束SOC")

            results.append({
                "source_order_no": source_order_no,
                "order_no": source_order_no,
                "business_order_no": _str(row, "业务订单号"),
                "order_type": _str(row, "订单类型"),
                "station_name": _str(row, "电站名称"),
                "channel": _str(row, "渠道来源"),
                "gun_code": _str(row, "枪编号"),
                "device_type": _str(row, "设备类型"),
                "user_code": _str(row, "用户编码"),
                "enterprise_name": _str(row, "企业名称"),
                "plate_number": _str(row, "车牌号"),
                "vin": _str(row, "车辆vin"),
                "status": status,
                "start_time": ChargingOrderImportService.parse_datetime(_str(row, "充电开始时间") or ""),
                "end_time": ChargingOrderImportService.parse_datetime(_str(row, "充电结束时间") or ""),
                "duration_minutes": _int(row, "充电时长(分)"),
                "charging_kwh": _float(row, "订单电量"),
                "original_amount": _float(row, "订单原价金额"),
                "energy_cost": _float(row, "电费"),
                "service_cost": _float(row, "服务费"),
                "pay_amount": _float(row, "订单实付金额"),
                "total_amount": _float(row, "订单实付金额"),
                "discount_amount": _float(row, "订单优惠金额"),
                "pay_method": _str(row, "支付方式"),
                "pay_status": pay_status,
                "start_soc": start_soc,
                "end_soc": end_soc,
                "stop_reason": _str(row, "停止原因"),
                "start_mode": _str(row, "启动方式"),
                "peak_kwh": _float(row, "峰电量"),
                "peak_cost": _float(row, "峰电费"),
                "flat_kwh": _float(row, "平电量"),
                "flat_cost": _float(row, "平电费"),
                "valley_kwh": _float(row, "谷电量"),
                "valley_cost": _float(row, "谷电费"),
                "sharp_kwh": _float(row, "尖电量"),
                "sharp_cost": _float(row, "尖电费"),
            })

        wb.close()
        return results

    @staticmethod
    def deduplicate(new_orders: list[dict], existing_refs: set[str]) -> list[dict]:
        seen = set()
        results = []
        for order in new_orders:
            ref = order.get("source_order_no")
            if not ref:
                continue
            if ref in existing_refs:
                continue
            if ref in seen:
                continue
            seen.add(ref)
            results.append(order)
        return results
