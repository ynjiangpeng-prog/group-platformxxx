import re
from datetime import date, datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.business.models import DailyExpense, WorkHourRecord
from app.models.charging.models import ChargingDevice, ChargingStation, FleetCustomer
from app.models.erp.models import Contract, Supplier
from app.models.finance.models import ArApRecord, Invoice
from app.models.erp.models import ProcurementRequest, PurchaseOrder
from app.models.organization import User
from app.models.project.models import Project, ConstructionLog

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
REQUIRED_FONT = Font(bold=True, color="FF0000", size=9)

TEMPLATE_REGISTRY = {
    "contract": {
        "model": Contract,
        "sheet_name": "合同导入",
        "headers": [
            "合同编号", "合同名称", "甲方", "乙方", "金额",
            "签订日期", "开始日期", "结束日期", "付款条件", "备注",
        ],
        "fields": [
            "contract_no", "name", "party_a", "party_b", "total_amount",
            "signing_date", "start_date", "end_date", "payment_terms", None,
        ],
        "widths": [20, 30, 25, 25, 15, 15, 15, 15, 20, 30],
        "required": ["contract_no", "name"],
        "date_fields": {"signing_date", "start_date", "end_date"},
        "number_fields": {"total_amount"},
        "int_fields": set(),
        "defaults": {"contract_type": "general", "status": "draft"},
    },
    "invoice": {
        "model": Invoice,
        "sheet_name": "发票导入",
        "headers": [
            "发票类型", "发票号码", "开票日期", "销方名称", "销方税号",
            "购方名称", "购方税号", "不含税金额", "税率", "税额", "价税合计",
        ],
        "fields": [
            "invoice_type", "invoice_no", "issue_date", "seller_name", None,
            "buyer_name", None, "amount_before_tax", "tax_rate", "tax_amount", "total_amount",
        ],
        "widths": [15, 20, 15, 25, 20, 25, 20, 15, 10, 15, 15],
        "required": ["invoice_type", "invoice_no"],
        "date_fields": {"issue_date"},
        "number_fields": {"amount_before_tax", "tax_rate", "tax_amount", "total_amount"},
        "int_fields": set(),
        "defaults": {"direction": "incoming", "check_status": "unchecked"},
    },
    "expense": {
        "model": DailyExpense,
        "sheet_name": "费用导入",
        "headers": ["日期", "分类", "金额", "付款方类型", "付款方", "说明", "项目ID"],
        "fields": ["expense_date", "category", "amount", "payer_type", "payer_name", "description", "project_id"],
        "widths": [15, 15, 15, 15, 15, 30, 36],
        "required": ["expense_date", "category", "amount", "payer_type"],
        "date_fields": {"expense_date"},
        "number_fields": {"amount"},
        "int_fields": set(),
        "defaults": {},
    },
    "project": {
        "model": Project,
        "sheet_name": "项目导入",
        "headers": [
            "项目编号", "项目名称", "项目类型", "优先级", "总预算",
            "开始日期", "结束日期", "省", "市", "地址",
        ],
        "fields": [
            "project_code", "name", "project_type", "priority", "total_budget",
            "start_date", "end_date", "province", "city", "address",
        ],
        "widths": [20, 30, 15, 10, 15, 15, 15, 10, 10, 30],
        "required": ["project_code", "name", "project_type"],
        "date_fields": {"start_date", "end_date"},
        "number_fields": {"total_budget"},
        "int_fields": {"priority"},
        "defaults": {"status": "draft"},
    },
    "station": {
        "model": ChargingStation,
        "sheet_name": "充电站导入",
        "headers": [
            "站点编号", "名称", "类型", "省", "市", "区",
            "地址", "经度", "纬度", "车位数", "月租金", "电力容量",
        ],
        "fields": [
            "station_code", "name", "station_type", "province", "city", "district",
            "address", "longitude", "latitude", "total_parking", "monthly_rent", "power_capacity",
        ],
        "widths": [20, 25, 10, 10, 10, 10, 30, 12, 12, 10, 12, 12],
        "required": ["station_code", "name"],
        "date_fields": set(),
        "number_fields": {"longitude", "latitude", "monthly_rent", "power_capacity"},
        "int_fields": {"total_parking"},
        "defaults": {"station_type": "public", "status": "planning"},
    },
    "supplier": {
        "model": Supplier,
        "sheet_name": "供应商导入",
        "headers": [
            "供应商编码", "供应商名称", "分类", "联系人", "联系电话",
            "统一信用代码", "开户银行", "银行账号", "评级", "备注",
        ],
        "fields": [
            "code", "name", "category", "contact_person", "contact_phone",
            "unified_credit_code", "bank_name", "bank_account", "rating", "remark",
        ],
        "widths": [20, 25, 15, 15, 15, 20, 20, 20, 10, 30],
        "required": ["code", "name", "category"],
        "date_fields": set(),
        "number_fields": set(),
        "int_fields": {"rating"},
        "defaults": {"status": 1, "rating": 3},
    },
    "work_hour": {
        "model": WorkHourRecord,
        "sheet_name": "工时导入",
        "headers": ["项目ID", "员工ID", "日期", "工时", "工作类型", "说明", "加班时长"],
        "fields": ["project_id", "employee_id", "work_date", "hours", "work_type", "description", "overtime_hours"],
        "widths": [36, 36, 15, 10, 15, 30, 10],
        "required": ["project_id", "employee_id", "work_date", "hours"],
        "date_fields": {"work_date"},
        "number_fields": {"hours", "overtime_hours"},
        "int_fields": set(),
        "defaults": {"status": "submitted"},
    },
    "fleet": {
        "model": FleetCustomer,
        "sheet_name": "车队客户导入",
        "headers": ["车队名称", "车队编码", "联系人", "联系电话", "车队规模", "信用额度", "备注"],
        "fields": ["fleet_name", "fleet_code", "contact_person", "contact_phone", "fleet_size", "credit_limit", "remark"],
        "widths": [25, 20, 15, 15, 10, 15, 30],
        "required": ["fleet_name", "fleet_code"],
        "date_fields": set(),
        "number_fields": {"credit_limit"},
        "int_fields": {"fleet_size"},
        "defaults": {"status": "active"},
    },
    "device": {
        "model": ChargingDevice,
        "sheet_name": "设备导入",
        "headers": ["站点ID", "设备编码", "制造商", "型号", "设备类型", "额定功率", "枪数量"],
        "fields": ["station_id", "device_code", "manufacturer", "model", "device_type", "rated_power", "gun_count"],
        "widths": [36, 20, 20, 20, 15, 12, 10],
        "required": ["station_id", "device_code"],
        "date_fields": set(),
        "number_fields": {"rated_power"},
        "int_fields": {"gun_count"},
        "defaults": {"device_type": "dc_fast", "status": "offline"},
    },
    "construction_log": {
        "model": ConstructionLog,
        "sheet_name": "施工日志导入",
        "headers": ["项目ID", "日期", "天气", "温度", "施工内容", "人数", "设备", "材料", "安全状态", "质量问题", "执行单位"],
        "fields": ["project_id", "log_date", "weather", "temperature", "work_content", "worker_count", "equipment_used", "materials_used", "safety_status", "quality_issues", "execution_unit"],
        "widths": [36, 15, 10, 10, 30, 8, 20, 20, 10, 20, 20],
        "required": ["project_id", "log_date"],
        "date_fields": {"log_date"},
        "number_fields": set(),
        "int_fields": {"worker_count"},
        "defaults": {"safety_status": "normal"},
    },
}

ENTITY_MODEL_MAP = {
    "contract": Contract,
    "invoice": Invoice,
    "expense": DailyExpense,
    "project": Project,
    "station": ChargingStation,
    "supplier": Supplier,
    "work_hour": WorkHourRecord,
    "fleet": FleetCustomer,
    "device": ChargingDevice,
    "procurement_request": ProcurementRequest,
    "purchase_order": PurchaseOrder,
    "ar_ap": ArApRecord,
}

APPROVE_CONFIG = {
    "invoice": {"field": "check_status", "value": "checked"},
    "ar_ap": {"field": "status", "value": "settled"},
    "expense": {"field": "status", "value": "approved"},
    "work_hour": {"field": "status", "value": "approved"},
}
DEFAULT_APPROVE = {"field": "status", "value": "approved"}


class ExcelTemplateGenerator:

    @staticmethod
    def _style_header(ws, headers, col_widths):
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 30
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"
        ws.freeze_panes = "A2"

    def generate(self, template_type: str) -> BytesIO:
        config = TEMPLATE_REGISTRY.get(template_type)
        if not config:
            raise ValueError(f"未知模板类型: {template_type}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = config["sheet_name"]
        self._style_header(ws, config["headers"], config["widths"])
        for col_idx, header in enumerate(config["headers"], 1):
            field = config["fields"][col_idx - 1]
            if field and field in config["required"]:
                cell = ws.cell(row=2, column=col_idx)
                cell.font = REQUIRED_FONT
                cell.value = "*必填"
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


class ExcelBatchImporter:

    @staticmethod
    def validate_headers(ws, expected_headers):
        actual = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, len(expected_headers) + 1)]
        return actual == expected_headers

    @staticmethod
    def _convert_value(value, field_name, config):
        if value is None:
            return None
        if field_name in config.get("date_fields", set()):
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if isinstance(value, str):
                value = value.strip()
                for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
                    try:
                        return datetime.strptime(value, fmt).date()
                    except ValueError:
                        continue
                raise ValueError(f"无法解析日期: {value}")
            return value
        if field_name in config.get("int_fields", set()):
            if isinstance(value, (int, float)):
                return int(value)
            if isinstance(value, str):
                cleaned = value.strip().replace(",", "")
                return int(float(cleaned))
            return int(value)
        if field_name in config.get("number_fields", set()):
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str):
                cleaned = value.strip().replace(",", "")
                return float(cleaned)
            return float(value)
        if isinstance(value, str):
            return value.strip()
        return value

    @staticmethod
    def parse_sheet(ws, config):
        expected_headers = config["headers"]
        fields = config["fields"]
        required = set(config["required"])
        valid_rows = []
        errors = []
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for row_idx, row in enumerate(rows, start=2):
            if not row or all(v is None for v in row):
                continue
            row_data = {}
            row_errors = []
            for col_idx, field in enumerate(fields):
                if col_idx >= len(row):
                    break
                raw = row[col_idx]
                if field is None:
                    continue
                if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                    if field in required:
                        row_errors.append({"row": row_idx, "field": config["headers"][col_idx], "message": f"{config['headers'][col_idx]}不能为空"})
                    continue
                try:
                    row_data[field] = ExcelBatchImporter._convert_value(raw, field, config)
                except (ValueError, TypeError) as e:
                    row_errors.append({"row": row_idx, "field": config["headers"][col_idx], "message": str(e)})
            if row_errors:
                errors.extend(row_errors)
                continue
            if not row_data:
                continue
            defaults = config.get("defaults", {})
            for k, v in defaults.items():
                row_data.setdefault(k, v)
            valid_rows.append(row_data)
        return valid_rows, errors

    async def _do_import(self, file_bytes, db: AsyncSession, user: User, template_type: str):
        config = TEMPLATE_REGISTRY.get(template_type)
        if not config:
            return {"total": 0, "success": 0, "failed": 0, "errors": [{"row": 0, "field": "", "message": f"未知模板类型: {template_type}"}]}

        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        ws = wb.active

        if not self.validate_headers(ws, config["headers"]):
            return {"total": 0, "success": 0, "failed": 0, "errors": [{"row": 0, "field": "", "message": "表头不匹配，请使用标准模板"}]}

        valid_rows, errors = self.parse_sheet(ws, config)
        total = len(valid_rows) + len(errors)
        success = 0
        model_cls = config["model"]

        for row_data in valid_rows:
            try:
                obj = model_cls(**row_data, company_id=user.company_id, created_by=user.id)
                db.add(obj)
                success += 1
            except Exception as e:
                errors.append({"row": 0, "field": "", "message": str(e)})

        if success > 0:
            try:
                await db.flush()
            except Exception as e:
                success = 0
                errors.insert(0, {"row": 0, "field": "", "message": f"数据库写入失败: {str(e)}"})

        return {
            "total": total,
            "success": success,
            "failed": len(errors),
            "errors": errors,
        }

    async def import_contracts(self, file_bytes, db, user):
        return await self._do_import(file_bytes, db, user, "contract")

    async def import_invoices(self, file_bytes, db, user):
        return await self._do_import(file_bytes, db, user, "invoice")

    async def import_expenses(self, file_bytes, db, user):
        return await self._do_import(file_bytes, db, user, "expense")

    async def import_projects(self, file_bytes, db, user):
        return await self._do_import(file_bytes, db, user, "project")

    async def import_stations(self, file_bytes, db, user):
        return await self._do_import(file_bytes, db, user, "station")

    async def import_suppliers(self, file_bytes, db, user):
        return await self._do_import(file_bytes, db, user, "supplier")

    async def import_work_hours(self, file_bytes, db, user):
        return await self._do_import(file_bytes, db, user, "work_hour")

    async def import_fleets(self, file_bytes, db, user):
        return await self._do_import(file_bytes, db, user, "fleet")

    async def import_devices(self, file_bytes, db, user):
        return await self._do_import(file_bytes, db, user, "device")

    async def import_construction_logs(self, file_bytes, db, user):
        return await self._do_import(file_bytes, db, user, "construction_log")

async def export_to_excel(db: AsyncSession, user: User, entity_type: str, filters: dict | None = None, columns: list[str] | None = None):
    config = TEMPLATE_REGISTRY.get(entity_type)
    if not config:
        raise ValueError(f"不支持导出的实体类型: {entity_type}")

    model_cls = config["model"]
    query = select(model_cls).where(
        model_cls.is_deleted == False,
        model_cls.company_id == user.company_id,
    )

    if filters:
        for key, value in filters.items():
            if hasattr(model_cls, key):
                col = getattr(model_cls, key)
                query = query.where(col == value)

    query = query.order_by(model_cls.created_at.desc()).limit(10000)
    result = await db.execute(query)
    items = result.scalars().all()

    if columns:
        col_indices = []
        for col_name in columns:
            if col_name in config["headers"]:
                col_indices.append(config["headers"].index(col_name))
            else:
                col_indices.append(None)
        export_headers = columns
        export_fields = []
        for idx in col_indices:
            export_fields.append(config["fields"][idx] if idx is not None else None)
    else:
        export_headers = config["headers"]
        export_fields = config["fields"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config["sheet_name"]
    ExcelTemplateGenerator._style_header(ws, export_headers, config["widths"])

    for row_idx, item in enumerate(items, 2):
        for col_idx, field in enumerate(export_fields, 1):
            if field is None:
                ws.cell(row=row_idx, column=col_idx, value="")
                continue
            val = getattr(item, field, None)
            if isinstance(val, date):
                val = val.isoformat()
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


template_gen = ExcelTemplateGenerator()
batch_importer = ExcelBatchImporter()
