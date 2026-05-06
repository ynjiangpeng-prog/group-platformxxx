import csv
import io
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation


class BankImportService:

    @staticmethod
    def parse_file(file_data: bytes, filename: str) -> list[dict]:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext == "csv":
            return BankImportService._parse_csv(file_data)
        if ext == "xls":
            return BankImportService._parse_xls(file_data)
        if ext == "xlsx":
            return BankImportService._parse_xlsx(file_data)
        raise ValueError(f"不支持的文件格式: .{ext}")

    @staticmethod
    def _parse_amount(text: str) -> float:
        if text is None:
            return 0.0
        text = str(text).strip().replace(",", "").replace(" ", "").replace("\u00a0", "")
        if not text or text == "-":
            return 0.0
        try:
            return float(Decimal(text))
        except (InvalidOperation, ValueError):
            return 0.0

    @staticmethod
    def _format_date(text: str) -> str | None:
        if not text:
            return None
        text = str(text).strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d",
            "%Y-%m-%d",
            "%Y%m%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%Y.%m.%d",
            "%Y.%m.%d %H:%M:%S",
        ):
            try:
                return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_csv(file_data: bytes) -> list[dict]:
        """Generic CSV parser for all banks: CMB, BOC, ICBC, etc.
        Auto-detects header row and column mapping."""
        for enc in ("utf-8-sig", "gb18030", "gbk", "gb2312"):
            try:
                text = file_data.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        else:
            text = file_data.decode("gb18030", errors="replace")

        lines = text.splitlines()
        if len(lines) < 3:
            return []

        # Column header matching rules: (column_key, [match_phrases])
        HEADER_RULES = [
            ("tx_date", ["交易日期", "记账日期", "日期", "转帐日期", "记账时间"]),
            ("tx_time", ["交易时间"]),
            ("tx_amount", ["交易金额", "发生额"]),
            ("debit", ["借方金额", "支出金额", "支出", "借方发生额", "借方"]),
            ("credit", ["贷方金额", "收入金额", "收入", "贷方发生额", "贷方"]),
            ("balance", ["余额", "账户余额", "当前余额", "交易后余额"]),
            ("payer", ["付款人名称", "付款户名"]),
            ("payee", ["收款人名称", "收款户名"]),
            ("payer_account", ["付款人账号"]),
            ("payee_account", ["收款人账号"]),
            ("payer_bank", ["付款人开户行名"]),
            ("payee_bank", ["收款人开户行名"]),
            ("tx_type_col", ["交易类型"]),
            ("source_ref", ["交易流水号", "流水号", "交易号", "凭证号", "唯一流水编号", "业务流水号", "记录标识号"]),
            ("summary", ["摘要"]),
            ("purpose", ["用途"]),
            ("remark", ["交易附言"]),
            ("account_name", ["户名", "账户名称"]),
            ("account_no", ["账号", "账户号码"]),
            ("status", ["交易状态", "状态", "交易结果", "处理状态"]),
        ]

        # Scan first 20 rows to find header row
        header_row_idx = None
        col_map = {}
        for row_idx, line in enumerate(lines[:20]):
            row_reader = csv.reader(io.StringIO(line))
            try:
                headers = next(row_reader)
            except StopIteration:
                continue
            mapped = {}
            for idx, h in enumerate(headers):
                h_stripped = str(h).strip()
                if not h_stripped:
                    continue
                for col_key, phrases in HEADER_RULES:
                    if col_key in mapped:
                        continue
                    for phrase in phrases:
                        if phrase in h_stripped:
                            mapped[col_key] = idx
                            break
            # Found header if we mapped at least tx_date + amount columns
            if "tx_date" in mapped and ("tx_amount" in mapped or "debit" in mapped or "credit" in mapped):
                header_row_idx = row_idx
                col_map = mapped
                break

        if header_row_idx is None:
            return []

        # Detect bank name from file content
        text_lower = text[:500].lower()
        bank_name = "银行"
        for bank_kw, bank_label in [
            ("招商银行", "招商银行"), ("中国银行", "中国银行"), ("工商银行", "工商银行"),
            ("建设银行", "建设银行"), ("农业银行", "农业银行"), ("交通银行", "交通银行"),
            ("兴业银行", "兴业银行"),
        ]:
            if bank_kw in text_lower:
                bank_name = bank_label
                break

        results = []
        for line in lines[header_row_idx + 1:]:
            if not line.strip():
                continue
            row_reader = csv.reader(io.StringIO(line))
            try:
                row = next(row_reader)
            except StopIteration:
                continue
            if len(row) <= max(col_map.values(), default=0):
                continue

            def cell(key):
                idx = col_map.get(key)
                if idx is None or idx >= len(row):
                    return ""
                return str(row[idx]).strip()

            # Filter by transaction status if present
            status = cell("status")
            if status and ("失败" in status or "撤销" in status or "退回" in status or "冲正" in status or "已拒绝" in status):
                continue

            # Parse date
            raw_date = cell("tx_date")
            tx_date = BankImportService._format_date(raw_date)
            if not tx_date:
                continue

            # Parse amount: prefer separate debit/credit, fall back to single amount
            debit = BankImportService._parse_amount(cell("debit"))
            credit = BankImportService._parse_amount(cell("credit"))
            if "debit" in col_map or "credit" in col_map:
                amount = credit - debit
            else:
                raw_amount = cell("tx_amount")
                sign = 1
                clean = raw_amount
                if clean.startswith("+"):
                    sign = 1
                    clean = clean[1:]
                elif clean.startswith("-"):
                    sign = -1
                    clean = clean[1:]
                amount = BankImportService._parse_amount(clean) * sign
            if amount == 0:
                continue

            balance = BankImportService._parse_amount(cell("balance")) if "balance" in col_map else None
            source_ref = cell("source_ref") or None
            summary = cell("summary") or None
            purpose = cell("purpose") or None
            remark_val = cell("remark") or None
            account_name_val = cell("account_name") or None
            account_no_val = cell("account_no") or None

            # Determine counterparty & account_name from payer/payee
            payer = cell("payer")
            payee = cell("payee")
            payer_account = cell("payer_account")
            payee_account = cell("payee_account")

            if amount >= 0:
                # Incoming: payer is the counterparty, payee is us
                counterparty = payer or None
                counterparty_account = payer_account or None
                if not account_name_val:
                    account_name_val = payee or None
                if not account_no_val:
                    account_no_val = payee_account or None
            else:
                # Outgoing: payee is the counterparty, payer is us
                counterparty = payee or None
                counterparty_account = payee_account or None
                if not account_name_val:
                    account_name_val = payer or None
                if not account_no_val:
                    account_no_val = payer_account or None

            results.append({
                "tx_date": tx_date,
                "tx_amount": round(amount, 2),
                "balance": round(balance, 2) if balance else None,
                "counterparty": counterparty,
                "counterparty_account": counterparty_account or None,
                "summary": summary or None,
                "purpose": purpose or None,
                "remark": remark_val or None,
                "tx_type": "income" if amount >= 0 else "expense",
                "source_ref": source_ref,
                "account_name": account_name_val,
                "account_no": account_no_val or None,
                "bank_name": bank_name,
            })

        return results

    @staticmethod
    def _parse_xls(file_data: bytes) -> list[dict]:
        import xlrd

        wb = xlrd.open_workbook(file_contents=file_data)
        sheet = wb.sheet_by_index(0)

        first_cell = sheet.cell_value(0, 0)

        first_cell_str = str(first_cell)
        second_cell_str = str(sheet.cell_value(1, 0)) if sheet.nrows > 1 else ""
        if first_cell_str.startswith("本表总笔数"):
            return BankImportService._parse_xls_yashida(sheet)
        elif "兴业银行交易明细" in first_cell_str:
            return BankImportService._parse_xls_cib_personal(sheet, wb)
        elif "云南省农村信用社" in first_cell_str or "按笔交易流水查询" in second_cell_str:
            return BankImportService._parse_xls_rural_commercial(sheet, wb)
        else:
            return BankImportService._parse_xls_cib_personal(sheet, wb)

    @staticmethod
    def _parse_xls_yashida(sheet) -> list[dict]:
        header_row = 1
        headers = [sheet.cell_value(header_row, c) for c in range(sheet.ncols)]

        col_map = {}
        for idx, h in enumerate(headers):
            h_stripped = str(h).strip()
            if "唯一流水编号" in h_stripped:
                col_map["source_ref"] = idx
            elif h_stripped == "户名":
                col_map["account_name"] = idx
            elif h_stripped == "账号":
                col_map["account_no"] = idx
            elif "借方金额" in h_stripped or "支出" in h_stripped:
                col_map["debit"] = idx
            elif "贷方金额" in h_stripped or "收入" in h_stripped:
                col_map["credit"] = idx
            elif "账户余额" in h_stripped or "余额" in h_stripped:
                col_map["balance"] = idx
            elif h_stripped == "摘要":
                col_map["summary"] = idx
            elif h_stripped == "用途":
                col_map["purpose"] = idx
            elif h_stripped == "备注":
                col_map["remark"] = idx
            elif "对方户名" in h_stripped:
                col_map["counterparty"] = idx
            elif "对方账号" in h_stripped:
                col_map["counterparty_account"] = idx
            elif "对方银行" in h_stripped:
                col_map["counterparty_bank"] = idx
            elif "交易日期" in h_stripped or "记账日期" in h_stripped or h_stripped == "日期":
                col_map["tx_date"] = idx

        results = []
        for r in range(header_row + 1, sheet.nrows):
            source_ref = str(sheet.cell_value(r, col_map.get("source_ref", 0))).strip()
            if not source_ref or source_ref == "":
                continue

            tx_date = None
            if "tx_date" in col_map:
                raw_val = sheet.cell_value(r, col_map["tx_date"])
                raw_date = str(raw_val).strip()
                cell_type = sheet.cell_type(r, col_map["tx_date"])
                if cell_type == 3:  # XL_CELL_DATE
                    try:
                        raw_date = xlrd.xldate_as_datetime(raw_val, wb.datemode).strftime("%Y-%m-%d")
                    except Exception:
                        pass
                tx_date = BankImportService._format_date(raw_date)
            if not tx_date:
                date_str = source_ref[:8] if len(source_ref) >= 8 and source_ref[:8].isdigit() else ""
                tx_date = BankImportService._format_date(date_str)

            debit = BankImportService._parse_amount(sheet.cell_value(r, col_map.get("debit", -1))) if "debit" in col_map else 0.0
            credit = BankImportService._parse_amount(sheet.cell_value(r, col_map.get("credit", -1))) if "credit" in col_map else 0.0
            amount = credit - debit

            balance = BankImportService._parse_amount(sheet.cell_value(r, col_map.get("balance", -1))) if "balance" in col_map else None
            summary = str(sheet.cell_value(r, col_map.get("summary", -1))).strip() if "summary" in col_map else None
            counterparty = str(sheet.cell_value(r, col_map.get("counterparty", -1))).strip() if "counterparty" in col_map else None
            counterparty_account = str(sheet.cell_value(r, col_map.get("counterparty_account", -1))).strip() if "counterparty_account" in col_map else None
            account_name = str(sheet.cell_value(r, col_map.get("account_name", -1))).strip() if "account_name" in col_map else None
            account_no = str(sheet.cell_value(r, col_map.get("account_no", -1))).strip() if "account_no" in col_map else None
            purpose = str(sheet.cell_value(r, col_map.get("purpose", -1))).strip() if "purpose" in col_map else None
            remark_val = str(sheet.cell_value(r, col_map.get("remark", -1))).strip() if "remark" in col_map else None

            results.append({
                "tx_date": tx_date,
                "tx_amount": round(amount, 2),
                "balance": round(balance, 2) if balance else None,
                "counterparty": counterparty or None,
                "counterparty_account": counterparty_account or None,
                "summary": summary or None,
                "purpose": purpose or None,
                "remark": remark_val or None,
                "tx_type": "income" if amount >= 0 else "expense",
                "source_ref": source_ref or None,
                "account_name": account_name or None,
                "account_no": account_no or None,
                "bank_name": "兴业银行",
            })

        return results

    @staticmethod
    def _parse_xls_cib_personal(sheet, wb=None) -> list[dict]:
        header_row = 10
        data_start = header_row + 1

        if sheet.nrows <= data_start:
            return []

        results = []
        for r in range(data_start, sheet.nrows):
            raw_val = sheet.cell_value(r, 0)
            cell_type = sheet.cell_type(r, 0)
            raw_date = str(raw_val).strip()
            if cell_type == 3 and wb:
                try:
                    raw_date = xlrd.xldate_as_datetime(raw_val, wb.datemode).strftime("%Y-%m-%d")
                except Exception:
                    pass
            if not raw_date or raw_date == "":
                continue

            tx_date = BankImportService._format_date(raw_date)

            expense_val = sheet.cell_value(r, 2)
            income_val = sheet.cell_value(r, 3)
            expense = BankImportService._parse_amount(expense_val) if expense_val else 0.0
            income = BankImportService._parse_amount(income_val) if income_val else 0.0

            if income > 0:
                amount = income
            elif expense > 0:
                amount = -expense
            else:
                continue

            balance = BankImportService._parse_amount(sheet.cell_value(r, 4))
            summary = str(sheet.cell_value(r, 5)).strip()
            counterparty = str(sheet.cell_value(r, 6)).strip()
            counterparty_bank = str(sheet.cell_value(r, 7)).strip()
            counterparty_account = str(sheet.cell_value(r, 8)).strip()

            results.append({
                "tx_date": tx_date,
                "tx_amount": round(amount, 2),
                "balance": round(balance, 2) if balance else None,
                "counterparty": counterparty or None,
                "counterparty_account": counterparty_account or None,
                "summary": summary or None,
                "purpose": summary or None,
                "remark": None,
                "tx_type": "income" if amount >= 0 else "expense",
                "source_ref": f"{tx_date}-{counterparty}-{abs(amount)}",
                "account_name": None,
                "account_no": None,
                "bank_name": "兴业银行",
            })

        return results

    @staticmethod
    def _parse_xlsx(file_data: bytes) -> list[dict]:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        sheet = wb.active

        is_wechat = False
        header_row_idx = None
        rows_iter = sheet.iter_rows(min_row=1, max_row=20, values_only=False)

        for row in rows_iter:
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                if "微信支付账单明细" in val:
                    is_wechat = True
                    break
            if is_wechat:
                break

        if not is_wechat:
            wb.close()
            return []

        header_row_idx = None
        rows_iter = sheet.iter_rows(min_row=1, max_row=25, values_only=True)
        for row_idx, row in enumerate(rows_iter, start=1):
            for cell_val in row:
                if cell_val and "交易时间" in str(cell_val):
                    header_row_idx = row_idx
                    break
            if header_row_idx:
                break

        if not header_row_idx:
            wb.close()
            return []

        results = []
        rows_iter = sheet.iter_rows(min_row=header_row_idx + 1, values_only=True)
        for row in rows_iter:
            if not row or len(row) < 9:
                continue

            raw_date = str(row[0]).strip() if row[0] else ""
            if not raw_date:
                continue

            direction = str(row[4]).strip() if row[4] else ""
            if not direction or direction == "/":
                continue

            tx_date = BankImportService._format_date(raw_date)
            counterparty = str(row[2]).strip() if row[2] else None
            raw_amount = str(row[5]).strip() if row[5] else "0"
            amount = BankImportService._parse_amount(raw_amount.replace("¥", ""))
            source_ref = str(row[8]).strip() if row[8] else None

            if direction == "收入":
                tx_type = "income"
            elif direction == "支出":
                tx_type = "expense"
                amount = -abs(amount)
            else:
                continue

            results.append({
                "tx_date": tx_date,
                "tx_amount": round(amount, 2),
                "balance": None,
                "counterparty": counterparty or None,
                "counterparty_account": None,
                "summary": None,
                "tx_type": tx_type,
                "source_ref": source_ref or None,
                "account_name": None,
                "account_no": None,
                "bank_name": "微信支付",
            })

        wb.close()
        return results

    @staticmethod
    def _parse_xls_rural_commercial(sheet, wb=None) -> list[dict]:
        import xlrd

        # Header row is at index 2 (3rd row)
        header_row = 2
        if sheet.nrows <= header_row + 1:
            return []

        headers = [str(sheet.cell_value(header_row, c)).strip() for c in range(sheet.ncols)]

        col_map = {}
        for idx, h in enumerate(headers):
            if h == "流水号":
                col_map["source_ref"] = idx
            elif h == "付款单位":
                col_map["payer"] = idx
            elif h == "付款人账号":
                col_map["payer_account"] = idx
            elif h == "金 额（元）" or "金额" in h:
                col_map["tx_amount"] = idx
            elif h == "收款单位名称":
                col_map["payee"] = idx
            elif h == "收款账号":
                col_map["payee_account"] = idx
            elif h == "用 途" or h == "用途":
                col_map["purpose"] = idx
            elif h == "交易类型":
                col_map["tx_type_name"] = idx
            elif h == "交易状态":
                col_map["status"] = idx
            elif h == "制单日期":
                col_map["tx_date"] = idx
            elif h == "付款单位开户机构":
                col_map["bank_name"] = idx

        results = []
        for r in range(header_row + 1, sheet.nrows):
            # Skip failed/cancelled transactions
            status = str(sheet.cell_value(r, col_map.get("status", -1))).strip() if "status" in col_map else ""
            if "失败" in status or "撤销" in status or "退回" in status:
                continue

            # Parse date
            raw_date = str(sheet.cell_value(r, col_map.get("tx_date", -1))).strip() if "tx_date" in col_map else ""
            if not raw_date:
                continue

            # Handle xlrd date type
            cell_type = sheet.cell_type(r, col_map["tx_date"]) if "tx_date" in col_map else -1
            if cell_type == 3 and wb:  # XL_CELL_DATE
                try:
                    raw_date = xlrd.xldate_as_datetime(sheet.cell_value(r, col_map["tx_date"]), wb.datemode).strftime("%Y-%m-%d")
                except Exception:
                    pass

            tx_date = BankImportService._format_date(raw_date)
            if not tx_date:
                continue

            # Parse amount (always expense from payer's perspective)
            amount = BankImportService._parse_amount(
                str(sheet.cell_value(r, col_map.get("tx_amount", -1))).strip() if "tx_amount" in col_map else "0"
            )
            if amount == 0:
                continue

            # This is payer's statement, so negative (expense)
            tx_amount = -abs(amount)

            payer = str(sheet.cell_value(r, col_map.get("payer", -1))).strip() if "payer" in col_map else ""
            payer_account = str(sheet.cell_value(r, col_map.get("payer_account", -1))).strip() if "payer_account" in col_map else ""
            payee = str(sheet.cell_value(r, col_map.get("payee", -1))).strip() if "payee" in col_map else ""
            payee_account = str(sheet.cell_value(r, col_map.get("payee_account", -1))).strip() if "payee_account" in col_map else ""
            purpose = str(sheet.cell_value(r, col_map.get("purpose", -1))).strip() if "purpose" in col_map else ""
            tx_type_name = str(sheet.cell_value(r, col_map.get("tx_type_name", -1))).strip() if "tx_type_name" in col_map else ""
            source_ref = str(sheet.cell_value(r, col_map.get("source_ref", -1))).strip() if "source_ref" in col_map else ""
            bank_name = str(sheet.cell_value(r, col_map.get("bank_name", -1))).strip() if "bank_name" in col_map else "云南农村商业银行"

            summary = purpose or tx_type_name or None

            results.append({
                "tx_date": tx_date,
                "tx_amount": round(tx_amount, 2),
                "balance": None,
                "counterparty": payee or None,
                "counterparty_account": payee_account or None,
                "summary": summary,
                "purpose": purpose or None,
                "tx_type": "expense",
                "source_ref": source_ref or None,
                "account_name": payer or None,
                "account_no": payer_account or None,
                "bank_name": bank_name,
            })

        return results

    @staticmethod
    def deduplicate(new_txs: list[dict], existing_refs: set[str]) -> list[dict]:
        seen = set()
        results = []
        for tx in new_txs:
            ref = tx.get("source_ref")
            if not ref:
                ref = f"{tx.get('tx_date', '')}-{tx.get('counterparty', '')}-{tx.get('tx_amount', '')}-{uuid.uuid4().hex[:8]}"
                tx["source_ref"] = ref

            if ref in existing_refs:
                continue
            if ref in seen:
                continue

            seen.add(ref)
            results.append(tx)

        return results
