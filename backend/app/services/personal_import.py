import io
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation

from app.services.bank_import import BankImportService


class PersonalImportService:

    @staticmethod
    def parse_file(file_data: bytes, filename: str) -> tuple[list[dict], str]:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext == "xlsx":
            return PersonalImportService._parse_xlsx(file_data)
        if ext == "xls":
            return PersonalImportService._parse_xls(file_data)
        raise ValueError(f"不支持的文件格式: .{ext}，仅支持 xlsx/xls")

    @staticmethod
    def _parse_amount(text) -> float:
        return BankImportService._parse_amount(text)

    @staticmethod
    def _format_date(text: str) -> str | None:
        return BankImportService._format_date(text)

    @staticmethod
    def _parse_xlsx(file_data: bytes) -> tuple[list[dict], str]:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        sheet = wb.active

        first_rows = []
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True)):
            first_rows.append(row)

        source_type = "unknown"
        for row in first_rows:
            for cell_val in row:
                val = str(cell_val) if cell_val else ""
                if "微信支付账单明细" in val:
                    source_type = "wechat"
                    break
                if "兴业银行" in val and ("交易明细" in val or "流水" in val):
                    source_type = "xingye_bank"
                    break
            if source_type != "unknown":
                break

        wb.close()

        if source_type == "wechat":
            wb = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
            results = PersonalImportService._parse_xlsx_wechat(wb.active)
            wb.close()
            return results, "wechat"

        if source_type == "xingye_bank":
            results = BankImportService._parse_xlsx(file_data)
            mapped = []
            for tx in results:
                mapped.append(PersonalImportService._to_personal(tx, "xingye_bank"))
            return mapped, "xingye_bank"

        raise ValueError("无法识别文件格式，请确认是招商银行/兴业银行或微信支付账单")

    @staticmethod
    def _parse_xls(file_data: bytes) -> tuple[list[dict], str]:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_data)
        sheet = wb.sheet_by_index(0)

        if PersonalImportService._is_cmb_xls(sheet):
            results = PersonalImportService._parse_xls_cmb(sheet, wb)
            wb.release_resources()
            return results, "cmb_bank"

        wb.release_resources()
        results = BankImportService._parse_xls(file_data)
        mapped = []
        for tx in results:
            bank = tx.get("bank_name", "")
            source = "xingye_bank" if "兴业" in bank else "manual"
            mapped.append(PersonalImportService._to_personal(tx, source))
        return mapped, "xingye_bank"

    @staticmethod
    def _is_cmb_xls(sheet) -> bool:
        for r in range(min(5, sheet.nrows)):
            row_vals = [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)]
            if "交易时间" in row_vals and "对方名称" in row_vals:
                return True
            joined = " ".join(row_vals)
            if "招商银行" in joined:
                return True
        return False

    @staticmethod
    def _parse_xls_cmb(sheet, wb) -> list[dict]:
        import xlrd

        header_row_idx = None
        col_map = {}
        for r in range(min(10, sheet.nrows)):
            headers = [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)]
            mapped = {}
            for idx, h in enumerate(headers):
                if h == "交易时间":
                    mapped["tx_time"] = idx
                elif h == "交易日期":
                    mapped["tx_date"] = idx
                elif h == "支出金额":
                    mapped["debit"] = idx
                elif h == "收入金额":
                    mapped["credit"] = idx
                elif h == "账户余额":
                    mapped["balance"] = idx
                elif h == "摘要":
                    mapped["summary"] = idx
                elif h == "交易对方":
                    mapped["counterparty_acct"] = idx
                elif h == "对方名称":
                    mapped["counterparty"] = idx
                elif h == "对方账号":
                    mapped["counterparty_account"] = idx
                elif h == "用途":
                    mapped["purpose"] = idx
                elif h == "商户单号":
                    mapped["merchant_ref"] = idx
                elif h == "备注":
                    mapped["remark"] = idx
            if "tx_date" in mapped or "tx_time" in mapped:
                header_row_idx = r
                col_map = mapped
                break

        if header_row_idx is None:
            return []

        results = []
        for r in range(header_row_idx + 1, sheet.nrows):
            def cell(key):
                idx = col_map.get(key)
                if idx is None or idx >= sheet.ncols:
                    return ""
                return str(sheet.cell_value(r, idx)).strip()

            tx_time = cell("tx_time")
            tx_date_raw = cell("tx_date")
            raw_date = tx_date_raw or tx_time
            if not raw_date:
                continue

            cell_type_date = sheet.cell_type(r, col_map["tx_date"]) if "tx_date" in col_map else -1
            if cell_type_date == 3:
                try:
                    raw_date = xlrd.xldate_as_datetime(
                        sheet.cell_value(r, col_map["tx_date"]), wb.datemode
                    ).strftime("%Y-%m-%d")
                except Exception:
                    pass
            else:
                if tx_time and tx_date_raw and ":" in tx_time:
                    raw_date = f"{tx_date_raw} {tx_time}"

            tx_date = PersonalImportService._format_date(raw_date)
            if not tx_date:
                continue

            debit = PersonalImportService._parse_amount(cell("debit"))
            credit = PersonalImportService._parse_amount(cell("credit"))
            if credit > 0:
                amount = credit
            elif debit > 0:
                amount = -debit
            else:
                continue

            balance = PersonalImportService._parse_amount(cell("balance")) if "balance" in col_map else None
            counterparty = cell("counterparty") or cell("counterparty_acct") or None
            counterparty_account = cell("counterparty_account") or None
            summary = cell("summary") or None
            purpose = cell("purpose") or None
            remark_val = cell("remark") or None
            merchant_ref = cell("merchant_ref") or None
            source_ref = merchant_ref or f"{tx_date}-{counterparty}-{abs(amount)}"

            results.append({
                "tx_date": tx_date,
                "tx_time": tx_time if tx_time and ":" in tx_time else None,
                "tx_amount": round(amount, 2),
                "balance": round(balance, 2) if balance else None,
                "counterparty": counterparty,
                "counterparty_name": counterparty,
                "counterparty_bank": None,
                "counterparty_account": counterparty_account or None,
                "description": summary or counterparty,
                "tx_type": "income" if amount >= 0 else "expense",
                "source": "cmb_bank",
                "source_account": "招商银行卡",
                "payment_channel": "bank_direct",
                "payment_method": "银行卡",
                "transaction_type": summary,
                "goods": None,
                "tx_status": None,
                "source_ref": source_ref,
                "fund_path": ["公户", "招商银行卡"],
                "is_public": True,
                "purpose": purpose,
                "remark": remark_val,
                "original_data": {
                    "tx_time": tx_time,
                    "summary": summary,
                    "purpose": purpose,
                    "remark": remark_val,
                    "counterparty_account": counterparty_account,
                    "merchant_ref": merchant_ref,
                },
            })

        return results

    @staticmethod
    def _parse_xlsx_wechat(sheet) -> list[dict]:
        header_row_idx = None
        rows_iter = sheet.iter_rows(min_row=1, max_row=25, values_only=True)
        for row_idx, row in enumerate(rows_iter, start=1):
            for cell_val in row:
                if cell_val and "交易时间" in str(cell_val):
                    header_row_idx = row_idx
                    break
            if header_row_idx:
                break

        if not header_row_idx:
            return []

        results = []
        rows_iter = sheet.iter_rows(min_row=header_row_idx + 1, values_only=True)
        for row in rows_iter:
            if not row or len(row) < 9:
                continue

            raw_date = str(row[0]).strip() if row[0] else ""
            if not raw_date:
                continue

            direction = str(row[4]).strip() if row[4] else ""
            if not direction or direction == "/":
                continue

            tx_date = PersonalImportService._format_date(raw_date)
            tx_time = None
            if " " in raw_date:
                parts = raw_date.split(" ", 1)
                tx_time = parts[1].strip() if len(parts) > 1 else None

            counterparty = str(row[2]).strip() if row[2] else None
            tx_type_val = str(row[1]).strip() if row[1] else None
            description = tx_type_val or counterparty
            raw_amount = str(row[5]).strip() if row[5] else "0"
            amount = PersonalImportService._parse_amount(raw_amount.replace("¥", ""))
            source_ref = str(row[8]).strip() if row[8] else None
            payment_method = str(row[6]).strip() if len(row) > 6 and row[6] else None
            goods = str(row[3]).strip() if len(row) > 3 and row[3] else None
            tx_status = str(row[7]).strip() if len(row) > 7 and row[7] else None

            if direction == "收入":
                tx_type = "income"
            elif direction == "支出":
                tx_type = "expense"
                amount = -abs(amount)
            else:
                continue

            results.append({
                "tx_date": tx_date,
                "tx_time": tx_time,
                "tx_amount": round(amount, 2),
                "balance": None,
                "counterparty": counterparty or None,
                "counterparty_name": counterparty,
                "counterparty_bank": None,
                "counterparty_account": None,
                "description": description or counterparty or None,
                "tx_type": tx_type,
                "source": "wechat",
                "source_account": "微信",
                "payment_channel": "wechat",
                "payment_method": payment_method,
                "transaction_type": tx_type_val,
                "goods": goods,
                "tx_status": tx_status,
                "source_ref": source_ref or None,
                "fund_path": ["公户", "兴业银行卡", "微信"],
                "is_public": True,
                "original_data": {
                    "direction": direction,
                    "payment_method": payment_method,
                    "counterparty": counterparty,
                    "description": description,
                    "goods": goods,
                    "tx_status": tx_status,
                    "transaction_type": tx_type_val,
                },
            })

        return results

    @staticmethod
    def _to_personal(tx: dict, source: str) -> dict:
        fund_path = ["公户", "招商银行卡"] if source == "cmb_bank" else (["公户", "兴业银行卡"] if source == "xingye_bank" else ["公户", "兴业银行卡", "微信"])
        return {
            "tx_date": tx.get("tx_date"),
            "tx_time": tx.get("tx_time"),
            "tx_amount": tx.get("tx_amount"),
            "balance": tx.get("balance"),
            "counterparty": tx.get("counterparty"),
            "counterparty_name": tx.get("counterparty_name") or tx.get("counterparty"),
            "counterparty_bank": tx.get("counterparty_bank"),
            "counterparty_account": tx.get("counterparty_account"),
            "description": tx.get("summary") or tx.get("counterparty"),
            "tx_type": tx.get("tx_type", "expense"),
            "source": source,
            "source_account": "招商银行卡" if source == "cmb_bank" else ("兴业银行卡" if source == "xingye_bank" else "微信"),
            "payment_channel": "bank_direct" if source in ("cmb_bank", "xingye_bank") else "wechat",
            "payment_method": tx.get("payment_method"),
            "transaction_type": tx.get("transaction_type"),
            "goods": tx.get("goods"),
            "tx_status": tx.get("tx_status"),
            "source_ref": tx.get("source_ref"),
            "fund_path": fund_path,
            "is_public": True,
            "purpose": tx.get("purpose"),
            "remark": tx.get("remark"),
            "original_data": tx,
        }

    @staticmethod
    def deduplicate(new_txs: list[dict], existing_refs: set[str]) -> list[dict]:
        seen = set()
        results = []
        for tx in new_txs:
            ref = tx.get("source_ref")
            if not ref:
                ref = f"{tx.get('tx_date', '')}-{tx.get('counterparty', '')}-{tx.get('tx_amount', '')}-{uuid.uuid4().hex[:8]}"
                tx["source_ref"] = ref

            if ref in existing_refs:
                continue
            if ref in seen:
                continue

            seen.add(ref)
            results.append(tx)

        return results
